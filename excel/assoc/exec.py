import re
from openpyxl import load_workbook, Workbook

def extract_course_info(description):
    course_codes = []
    course_types = []
    pattern = r'(?<=/courses/)[a-zA-Z0-9]+(?=.html)'
    matches = re.findall(pattern, description)
    course_codes.extend(matches)

    # Check for course type (exit-only or articulated)
    if 'articulated' in description.lower():
        course_types.append('Articulated')
    elif 'exit-only' in description.lower():
        course_types.append('Exit-Only')
    else:
        course_types.append('Unknown')  # Default type if not specified

    return course_codes, course_types

def main():
    # Load the workbook
    wb = load_workbook('Associations.xlsx')  # Replace 'your_excel_file.xlsx' with your file path
    ws = wb['Ticking Off Sheet']

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = 'out'

    rows = ws.max_row
    cell_intptr = 1

    # Iterate over each row in column G
    for r in range(1, rows):
        description = ws['G'+str(r)].value
        if (not description or (len(description) == 0)):
            continue

        # Extract course codes and course types from the description
        course_codes, course_types = extract_course_info(description)

        # Populate course codes and course types into specific columns and rows
        owning_code = ws['A'+str(r)].value
        owning_course = ws['C'+str(r)].value
        if (owning_code == 'C04430'):
            print (F"{course_codes}")
        for course_code in course_codes:
            out_ws.cell(row=cell_intptr, column=1).value = owning_course
            out_ws.cell(row=cell_intptr, column=2).value = owning_code
            out_ws.cell(row=cell_intptr, column=3).value = course_code
            out_ws.cell(row=cell_intptr, column=5).value = description
            cell_intptr += 1
        out_ws.cell(row=cell_intptr, column=4).value = course_types[0] if (len(course_types) > 0) else 'Unknown'

    out_wb.save ('out.xlsx')

if __name__ == "__main__":
    main()
