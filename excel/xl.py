import re
from openpyxl import load_workbook

def extract_course_codes(description):
    if (not description):
        return []
    if (len (description)== 0):
        return []
    course_codes = []
    pattern = r'(?<=/courses/)[a-zA-Z0-9]+(?=.html)'
    matches = re.findall(pattern, description)
    print (matches)
    course_codes.extend(matches)
    return course_codes

def main():
    # Load the workbook
    wb = load_workbook('Associations.xlsx')  # Replace 'your_excel_file.xlsx' with your file path
    ws = wb['Ticking Off Sheet']

    # Iterate over each row in column G
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=7, max_col=7, values_only=True):  # Column G is column 7
        description = row[0]  # Assuming only one cell in the row contains the description
        if (not description or len (description) == 0):
            continue

        # Extract course codes from the description
        course_codes = extract_course_codes(description)

        # Populate course codes into specific columns and rows
        start_row = ws.max_row + 1  # Start from the next available row
        start_column = 2  # Start from the second column (B)
        for course_code in course_codes:
            ws.cell(row=start_row, column=start_column).value = course_code
            start_row += 1

    # Save the workbook
    wb.save('better.xlsx')  # Replace 'your_updated_excel_file.xlsx' with your desired output file path

if __name__ == "__main__":
    main()

